import os
import tempfile
from openpyxl import load_workbook
from battery_optimizer.optimizer import solve_lp
from battery_optimizer.results import process_all_results
from battery_optimizer.report import export_results_to_excel
from battery_optimizer.default_data import (
    get_default_collection_centers,
    get_default_recycling_facilities,
    get_default_transport_costs,
)


def get_report_data():
    cc = get_default_collection_centers()
    rf = get_default_recycling_facilities()
    tc = get_default_transport_costs()
    opt_result = solve_lp(cc, rf, tc)
    processed = process_all_results(opt_result, cc, rf, tc)
    return opt_result, processed, cc, rf, tc


def test_export_results_to_excel_creates_file():
    opt_result, processed, cc, rf, tc = get_report_data()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "optimization_results.xlsx")
        result_path = export_results_to_excel(opt_result, processed, cc, rf, tc, path)

        assert os.path.exists(result_path)


def test_export_results_to_excel_has_required_sheets():
    opt_result, processed, cc, rf, tc = get_report_data()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "optimization_results.xlsx")
        result_path = export_results_to_excel(opt_result, processed, cc, rf, tc, path)

        workbook = load_workbook(result_path, data_only=False)
        sheet_names = workbook.sheetnames
        workbook.close()

        expected_sheets = [
            "diagnostics_summary",
            "summary",
            "allocation_matrix",
            "route_allocation",
            "facility_utilization",
            "supply_usage",
            "constraints",
            "interpretation",
        ]

        for sheet in expected_sheets:
            assert sheet in sheet_names


def test_diagnostics_summary_sheet_has_diagnostic_fields():
    opt_result, processed, cc, rf, tc = get_report_data()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "optimization_results.xlsx")
        result_path = export_results_to_excel(opt_result, processed, cc, rf, tc, path)

        workbook = load_workbook(result_path, data_only=True)
        worksheet = workbook["diagnostics_summary"]
        metrics = [worksheet.cell(row=row, column=1).value for row in range(2, worksheet.max_row + 1)]
        workbook.close()

        expected_metrics = [
            "Supply Absorption (%)",
            "System Capacity Utilization (%)",
            "Maximum Facility Utilization (%)",
            "Binding Capacity Facilities",
            "Binding Supply Centers",
        ]

        for metric in expected_metrics:
            assert metric in metrics


def test_summary_sheet_has_economic_metrics():
    opt_result, processed, cc, rf, tc = get_report_data()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "optimization_results.xlsx")
        result_path = export_results_to_excel(opt_result, processed, cc, rf, tc, path)

        workbook = load_workbook(result_path, data_only=True)
        worksheet = workbook["summary"]
        metrics = [worksheet.cell(row=row, column=1).value for row in range(2, worksheet.max_row + 1)]
        workbook.close()

        assert "Solver Status" in metrics
        assert "Objective Status" in metrics
        assert "Minimum Net Cost Objective (Rp/year)" in metrics


def test_route_allocation_sheet_has_user_friendly_headers():
    opt_result, processed, cc, rf, tc = get_report_data()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "optimization_results.xlsx")
        result_path = export_results_to_excel(opt_result, processed, cc, rf, tc, path)

        workbook = load_workbook(result_path, data_only=True)
        worksheet = workbook["route_allocation"]
        headers = [cell.value for cell in worksheet[1]]
        workbook.close()

        expected_headers = [
            "Collection Center ID",
            "Collection Center",
            "Recycling Facility ID",
            "Recycling Facility",
            "Transport Cost (Rp/kg)",
            "Processing Cost (Rp/kg)",
            "Recovery Revenue (Rp/kg)",
            "Net Cost Coefficient (Rp/kg)",
            "Allocated Volume (kg/year)",
            "Net Cost Objective (Rp/year)",
            "Route Net Benefit (Rp/year)",
        ]

        for header in expected_headers:
            assert header in headers


def test_rupiah_columns_remain_numeric():
    opt_result, processed, cc, rf, tc = get_report_data()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "optimization_results.xlsx")
        result_path = export_results_to_excel(opt_result, processed, cc, rf, tc, path)

        workbook = load_workbook(result_path, data_only=False)
        worksheet = workbook["route_allocation"]
        headers = [cell.value for cell in worksheet[1]]
        column_index = headers.index("Route Net Benefit (Rp/year)") + 1
        value = worksheet.cell(row=2, column=column_index).value
        number_format = worksheet.cell(row=2, column=column_index).number_format
        workbook.close()

        assert isinstance(value, (int, float))
        assert "Rp" in number_format


def test_interpretation_sheet_has_text():
    opt_result, processed, cc, rf, tc = get_report_data()

    with tempfile.TemporaryDirectory() as tmpdir:
        path = os.path.join(tmpdir, "optimization_results.xlsx")
        result_path = export_results_to_excel(opt_result, processed, cc, rf, tc, path)

        workbook = load_workbook(result_path, data_only=True)
        worksheet = workbook["interpretation"]
        text = worksheet.cell(row=2, column=1).value
        workbook.close()

        assert isinstance(text, str)
        assert len(text) > 50