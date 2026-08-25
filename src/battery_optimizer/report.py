import os
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from battery_optimizer.config import RESULTS_PATH
from battery_optimizer.results import build_economic_summary


RUPIAH_FORMAT = '[$Rp-421] #,##0.00'
NUMBER_FORMAT = '#,##0.00'
INTEGER_FORMAT = '#,##0'
SECONDS_FORMAT = '0.0000'
PERCENT_FORMAT = '0.00'


ALLOCATION_MATRIX_LABELS = {
    "cc_id": "Collection Center ID",
    "total_allocated_kg": "Total Allocated (kg/year)",
}

ROUTE_TABLE_LABELS = {
    "cc_id": "Collection Center ID",
    "cc_name": "Collection Center",
    "rf_id": "Recycling Facility ID",
    "rf_name": "Recycling Facility",
    "transport_cost_rp_kg": "Transport Cost (Rp/kg)",
    "processing_cost_rp_kg": "Processing Cost (Rp/kg)",
    "recovery_revenue_rp_kg": "Recovery Revenue (Rp/kg)",
    "net_cost_coeff_rp_kg": "Net Cost Coefficient (Rp/kg)",
    "allocated_kg": "Allocated Volume (kg/year)",
    "route_cost_rp": "Net Cost Objective (Rp/year)",
    "route_economic_benefit_rp": "Route Net Benefit (Rp/year)",
}

FACILITY_UTILIZATION_LABELS = {
    "rf_id": "Recycling Facility ID",
    "name": "Recycling Facility",
    "capacity_kg": "Capacity (kg/year)",
    "allocated_kg": "Allocated Volume (kg/year)",
    "unused_capacity_kg": "Unused Capacity (kg/year)",
    "utilization_pct": "Utilization (%)",
}

SUPPLY_USAGE_LABELS = {
    "cc_id": "Collection Center ID",
    "name": "Collection Center",
    "supply_kg": "Supply (kg/year)",
    "allocated_kg": "Allocated Volume (kg/year)",
    "unused_kg": "Unused Supply (kg/year)",
    "usage_pct": "Supply Usage (%)",
}

CONSTRAINT_TABLE_LABELS = {
    "constraint_type": "Constraint Type",
    "id": "Constraint ID",
    "name": "Name",
    "rhs_kg": "Limit (kg/year)",
    "slack_kg": "Slack (kg/year)",
    "shadow_price_rp_kg": "Shadow Price (Rp/kg)",
    "binding": "Binding Status",
}


def apply_header_style(worksheet):
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    worksheet.freeze_panes = "A2"


def set_column_widths(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)


def get_header_map(worksheet):
    return {
        cell.value: cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }


def apply_format_to_columns(worksheet, column_names, number_format):
    header_map = get_header_map(worksheet)
    for column_name in column_names:
        if column_name in header_map:
            column_index = header_map[column_name]
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format


def apply_summary_format(worksheet):
    for row_index in range(2, worksheet.max_row + 1):
        metric = worksheet.cell(row=row_index, column=1).value
        value_cell = worksheet.cell(row=row_index, column=2)
        if not isinstance(value_cell.value, (int, float)):
            continue
        metric_text = str(metric)
        if "Rp/year" in metric_text:
            value_cell.number_format = RUPIAH_FORMAT
        elif "(kg)" in metric_text:
            value_cell.number_format = NUMBER_FORMAT
        elif "(%)" in metric_text:
            value_cell.number_format = PERCENT_FORMAT
        elif "seconds" in metric_text:
            value_cell.number_format = SECONDS_FORMAT


def prepare_allocation_matrix(processed):
    df = processed["allocation_matrix"].copy()
    df.reset_index(inplace=True)
    return df.rename(columns=ALLOCATION_MATRIX_LABELS)


def prepare_route_table(processed):
    df = processed["route_table"].copy()
    return df.rename(columns=ROUTE_TABLE_LABELS)


def prepare_facility_utilization(processed):
    df = processed["facility_utilization"].copy()
    return df.rename(columns=FACILITY_UTILIZATION_LABELS)


def prepare_supply_usage(processed):
    df = processed["supply_usage"].copy()
    return df.rename(columns=SUPPLY_USAGE_LABELS)


def prepare_constraint_table(processed):
    df = processed["constraint_table"].copy()
    if "binding" in df.columns:
        df["binding"] = df["binding"].map(
            lambda value: "Binding" if value is True else ("Not Binding" if value is False else "N/A")
        )
    return df.rename(columns=CONSTRAINT_TABLE_LABELS)


def prepare_diagnostics_summary(processed):
    metrics = processed.get("system_metrics", {})
    rows = [
        ["Supply Absorption (%)", metrics.get("supply_absorption_pct")],
        ["System Capacity Utilization (%)", metrics.get("system_capacity_utilization_pct")],
        ["Maximum Facility Utilization (%)", metrics.get("max_facility_utilization_pct")],
        ["Most Utilized Facility", metrics.get("most_utilized_facility", "N/A")],
        ["Binding Capacity Facilities", ", ".join(metrics.get("binding_capacity_facilities", []))],
        ["Binding Supply Centers", ", ".join(metrics.get("binding_supply_centers", []))],
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def export_results_to_excel(opt_result, processed, cc_df, rf_df, tc_df, filepath=None):
    if filepath is None:
        filepath = RESULTS_PATH

    directory = os.path.dirname(filepath)
    if directory:
        os.makedirs(directory, exist_ok=True)

    economic_summary = processed.get("economic_summary", build_economic_summary(opt_result))
    diagnostics_summary = prepare_diagnostics_summary(processed)

    total_supply = float(cc_df["supply_kg"].sum())
    total_capacity = float(rf_df["capacity_kg"].sum())
    total_allocated = float(processed["supply_usage"]["allocated_kg"].sum())
    total_unused_supply = float(processed["supply_usage"]["unused_kg"].sum())
    total_unused_capacity = float(processed["facility_utilization"]["unused_capacity_kg"].sum())

    summary_rows = [
        ["Solver Status", opt_result["status"]],
        ["Objective Status", economic_summary["economic_status"]],
        ["Minimum Net Cost Objective (Rp/year)", opt_result["objective_value"]],
        [f"{economic_summary['economic_value_label']} (Rp/year)", economic_summary["net_economic_value_abs"]],
        ["Runtime (seconds)", round(opt_result["runtime"], 4)],
        ["Number of Collection Centers", len(cc_df)],
        ["Number of Recycling Facilities", len(rf_df)],
        ["Total Supply (kg)", total_supply],
        ["Total Capacity (kg)", total_capacity],
        ["Total Allocated (kg)", total_allocated],
        ["Unused Supply (kg)", total_unused_supply],
        ["Unused Capacity (kg)", total_unused_capacity],
        ["Objective Interpretation", economic_summary["economic_message"]],
    ]

    summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])
    alloc_matrix = prepare_allocation_matrix(processed)
    route_table = prepare_route_table(processed)
    facility_util = prepare_facility_utilization(processed)
    supply_usage = prepare_supply_usage(processed)
    constraint_table = prepare_constraint_table(processed)

    interp_df = pd.DataFrame(
        {"Optimization Interpretation": [processed["interpretation"]]}
    )

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        diagnostics_summary.to_excel(writer, sheet_name="diagnostics_summary", index=False)
        summary_df.to_excel(writer, sheet_name="summary", index=False)
        alloc_matrix.to_excel(writer, sheet_name="allocation_matrix", index=False)
        route_table.to_excel(writer, sheet_name="route_allocation", index=False)
        facility_util.to_excel(writer, sheet_name="facility_utilization", index=False)
        supply_usage.to_excel(writer, sheet_name="supply_usage", index=False)
        constraint_table.to_excel(writer, sheet_name="constraints", index=False)
        interp_df.to_excel(writer, sheet_name="interpretation", index=False)

        for worksheet in writer.book.worksheets:
            apply_header_style(worksheet)
            set_column_widths(worksheet)

        apply_summary_format(writer.sheets["diagnostics_summary"])
        apply_summary_format(writer.sheets["summary"])

        apply_format_to_columns(
            writer.sheets["allocation_matrix"],
            [column for column in alloc_matrix.columns if column != "Collection Center ID"],
            NUMBER_FORMAT,
        )

        apply_format_to_columns(
            writer.sheets["route_allocation"],
            [
                "Transport Cost (Rp/kg)",
                "Processing Cost (Rp/kg)",
                "Recovery Revenue (Rp/kg)",
                "Net Cost Coefficient (Rp/kg)",
                "Net Cost Objective (Rp/year)",
                "Route Net Benefit (Rp/year)",
            ],
            RUPIAH_FORMAT,
        )

        apply_format_to_columns(
            writer.sheets["route_allocation"],
            ["Allocated Volume (kg/year)"],
            NUMBER_FORMAT,
        )

        apply_format_to_columns(
            writer.sheets["facility_utilization"],
            [
                "Capacity (kg/year)",
                "Allocated Volume (kg/year)",
                "Unused Capacity (kg/year)",
            ],
            NUMBER_FORMAT,
        )

        apply_format_to_columns(
            writer.sheets["facility_utilization"],
            ["Utilization (%)"],
            PERCENT_FORMAT,
        )

        apply_format_to_columns(
            writer.sheets["supply_usage"],
            [
                "Supply (kg/year)",
                "Allocated Volume (kg/year)",
                "Unused Supply (kg/year)",
            ],
            NUMBER_FORMAT,
        )

        apply_format_to_columns(
            writer.sheets["supply_usage"],
            ["Supply Usage (%)"],
            PERCENT_FORMAT,
        )

        apply_format_to_columns(
            writer.sheets["constraints"],
            [
                "Limit (kg/year)",
                "Slack (kg/year)",
            ],
            NUMBER_FORMAT,
        )

        apply_format_to_columns(
            writer.sheets["constraints"],
            ["Shadow Price (Rp/kg)"],
            RUPIAH_FORMAT,
        )

    return filepath